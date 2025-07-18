# ===================================================================================
#       UPDATE MASSAL MARKETPLACE - TIKTOK & SHOPEE (GUI VERSION)
# ===================================================================================
# Versi: 2.6 (Solusi Ikon Final)
# Tanggal: 17 Juni 2024
# ===================================================================================

import tkinter as tk
from tkinter import ttk, filedialog, scrolledtext, messagebox, PhotoImage
import pandas as pd
import numpy as np
import os
import sys
import time
import warnings
import openpyxl
import threading
from queue import Queue
import webbrowser
import base64
from io import BytesIO
import ctypes
from ctypes import wintypes

# Mengabaikan UserWarning dari openpyxl
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# --- FUNGSI PENTING UNTUK MENEMUKAN FILE SETELAH DIKOMPILASI ---
def resource_path(relative_path):
    """ Dapatkan path absolut ke sumber daya, berfungsi untuk dev dan PyInstaller """
    try:
        # PyInstaller membuat folder temp _MEIPASS dan menyimpan path-nya di sys
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

# --- DATA LOGO (Base64 encoded logo.png) untuk header GUI ---
LOGO_BASE64 = """
iVBORw0KGgoAAAANSUhEUgAAAaAAAAFgCAYAAAC8frGRAAAAAXNSR0IArs4c6QAAAARnQU1BAACx
jwv8YQUAAAAJcEhZcwAADsMAAA7DAcdvqGgAAP+6SURBVHictP9p3FvXted7n/PudG9N773v
fTe7u/c1n3u9c961e93t6u7uzn3Nbrd7n7f3pvelhRIgQZIkSZAkyZIkSZLkyZwkxXLeJUlS
pA6EIEkyBEgSJkGSwEkgSV4nSZIkSZIkSZIkSZIkSVL9R4JkSZIkSZIkSZIkSZIkSZLUL5Eg
SZIkyZIkSZIkSZIkSVK/RIIkSZIkyZIkSZIkSZIkSfpFJEiSJEkyBEgCJEkyjAQSJEkyxJAg
SZIkyZIk9SskSJIkyZIkSZIkSZIkSVL/RIIkSZIkyZIkSZIkSZIkSfpFJEiSJEkyJEiCJEky
jAQSJEkyxJAgSZIkyZIk9SskSJIkyZIkSZIkSZIkSVL/RIIkSZIkyZIkSZIkSZIkSfpFJEiS
JEkyJEiCJEkyjAQSJEkyxJAgSZIkyZIk9SskSJIkyZIkSZIkSZIkSVL/RIIkSZIkyZIkSZIk
SZIkSfpFJEiSJEkyJEiCJEkyjAQSJEkyxJAgSZIkyZIk9SskSJIkyZIkSZIkSZIkSVL/RIIk
SZIkyZIkSZIkSZIkSfpFJEiSJEkyJEiCJEkyjAQSJEkyxJAgSZIkyZIk9SskSJIkyZIkSZIk
SZIkSVL/RIIkSZIkyZIkSZIkSZIkSfpFJEiSJEkyJEiCJEkyjAQSJEkyxJAgSZIkyZIk9Ssk
SJIkyZIkSZIkSZIkSVL/RIIkSZIkyZIkSZIkSZIkSfpFJEiSJEkyJEiCJEkyjAQSJEkyxJAg
SZIkyZIk9SskSJIkyZIkSZIkSZIkSVL/RIIkSZIkyZIkSZIkSZIkSfpFJEiSJEkyJEiCJEky
jAQSJEkyxJAgSZIkyZIk9SskSJIkyZIkSZIkSZIkSVL/RIIkSZIkyZIkSZIkSZIkSfpFJEiS
JEkyJEiCJEkyjAQSJEkyxJAgSZIkyZIk9SskSJIkyZIkSZIkSZIkSVL/RIIkSZIkyZIkSZIk
SZIkSfpFJEiSJEkyJEiCJEkyjAQSJEkyxJAgSZIkyZIk9SskSJIkyZIkSZIkSZIkSVL/RIIk
SZIkyZIkSZIkSZIkSfpFJEiSJEkyJEiCJEkyjAQSJEkyxJAgSZIkyZIk9SskSJIkyZIkSZIk
SZIkSVL/RIIkSZIkyZIkSZIkSZIkSfpFJEiSJEkyJEiCJEkyjAQSJEkyxJAgSZIkyZIk9Ssk
SJIkyZIkSZIkSZIkSVL/RIIkSZIkyZIkSZIkSZIkSfpFJEiSJEkyJEiCJEkyjAQSJEkyxJAg
SZIkyZIk9SskSJIkyZIkSZIkSZIkSVL/RIIkSZIkyZIkSZIkSZIkSfpFJEiSJEkyJEiCJEky
jAQSJEkyxJAgSZIkyZIk9SskSJIkyZIkSZIkSZIkSVL/RIIkSZIkyZIkSZIkSZIkSfpFJEiS
JEkyJEiCJEkyjAQSJEkyxJAgSZIkyZIk9SskSJIkyZIkSZIkSZIkSVL/RIIkSZIkyZIkSZIk
SZIkSfpFJEiSJEkyJEiCJEkyjAQSJEkyxJAgSZIkyZIk9SskSJIkyZIkSZIkSZIkSVL/RIIk
SZIkyZIkSZIkSZIkSfpFJEiSJEkyJEiCJEkyjAQSJEkyxJAgSZIkyZIk9SskSJIkyZIkSZIk
SZIkSVL/RIIkSZIkyZIkSZIkSZIkSfpFJEiSJEkyJEiCJEkyjAQSJEkyxJAgSZIkyZIk9Ssk
SJIkyZIkSZIkSZIkSVL/RIIkSZIkyZIkSZIkSZIkSfpFJEiSJEkyJEiCJEkyjAQSJEkyxJAg
SZIkyZIk9SskSJIkyZIkSZIkSZIkSVL/RIIkSZIkyZIkSZIkSZIkSfpFJEiSJEkyJEiCJEky
jAQSJEkyxJAgSZIkyZIk9SskSJIkyZIkSZIkSZIkSVL/RIIkSZIkyZIkSZIkSZIkSfpFJEiS
JEkyJEiCJEkyjAQSJEkyxJAgSZIkyZIk9SskSJIkyZIkSZIkSZIkSVL/RIIkSZIkyZIkSZIk
SZIkSfpFJEiSJEkyJEiCJEkyjAQSJEkyxJAgSZIkyZIk9SskSJIkyZIkSZIkSZIkSVL/RIIk
SZIkyZIkSZIkSZIkSfpFJEiSJEkyJEiCJEkyjAQSJEkyxJAgSZIkyZIk9SskSJIkyZIkSZIk
SZIkSVL/RIIkSZIkyZIkSZIkSZIkSfpFJEiSJEkyJEiCJEkyjAQSJEkyxJAgSZIkyZIk9Ssk
SJIkyZIkSZIkSZIkSVL/RIIkSZIkyZIkSZIkSZIkSfpFJEiSJEkyJEiCJEkyjAQSJEkyxJAg
SZIkyZIk9SskSJIkyZIkSZIkSZIkSVL/RIIkSZIkyZIkSZIkSZIkSfpFJEiSJEkyJEiCJEky
jAQSJEkyxJAgSZIkyZIk9SskSJIkyZIkSZIkSZIkSVL/RIIkSZIkyZIkSZIkSZIkSfpFJEiS
JEkyJEiCJEkyjAQSJEkyxJAgSZIkyZIk9SskSJIkyZIkSZIkSZIkSVL/RIIkSZIkyZIkSZIk
SZIkSfpFJEiSJEkyJEiCJEkyjAQSJEkyxJAgSZIkyZIk9SskSJIkyZIkSZIkSZIkSVL/RIIk
SZIkyZIkSZIkSZIkSfpFJEiSJEkyJEiCJEkyjAQSJEkyxJAgSZIkyZIk9SskSJIkyZIkSZIk
SZIkSVL/RIIkSZIkyZIkSZIkSZIkSfpFJEiSJEkyJEiCJEkyjAQSJEkyxJAgSZIkyZIk9Ssk
SJIkyZIkSZIkSZIkSVL/RIIkSZIkyZIkSZIkSZIkSfpFJEiSJEkyJEiCJEkyjAQSJEkyxJAg
SZIkyZIk9SskSJIkyZIkSZIkSZIkSVL/RIIkSZIkyZIkSZIkSZIkSfpFJEiSJEkyJEiCJEky
jAQSJEkyxJAgSZIkyZIk9SskSJIkyZIkSZIkSZIkSVL/RIIkSZIkyZIkSZIkSZIkSfpFJEiS
JEkyJEiCJEkyjAQSJEkyxJAgSZIkyZIk9SskSJIkyZIkSZIkSZIkSVL/RIIkSZIkyZIkSZIk
SZIkSfpFJEiSJEkyJEiCJEkyjAQSJEkyxJAgSZIkyZIk9SskSJIkyZIkSZIkSZIkSVL/RIIk
SZIkyZIkSZIkSZIkSfpFJEiSJEkyJEiCJEkyjAQSJEkyxJAgSZIkyZIk9SskSJIkyZIkSZIk
SZIkSVL/RIIkSZIkyZIkSZIkSZIkSfpFJEiSJEkyJEiCJEkyjAQSJEkyxJAgSZIkyZIk9Ssk
SJIkyZIkSZIkSZIkSVL/RIIkSZIkyZIkSZIkSZIkSfpFJEiSJEkyJEiCJEkyjAQSJEkyxJAg
SZIkyZIk9SskSJIkyZIkSZIkSZIkSVL/RIIkSZIkyZIkSZIkSZIkSfpFJEiSJEkyJEiCJEky
jAQSJEkyxJAgSZIkyZIk9SskSJIkyZIkSZIkSZIkSVL/RIIkSZIkyZIkSZIkSZIkSfpFJEiS
JEkyJEiCJEkyjAQSJEkyxJAgSZIkyZIk9SskSJIkyZIkSZIkSZIkSVL/RIIkSZIkyZIkSZIk
SZIkSfpFJEiSJEkyJEiCJEkyjAQSJEkyxJAgSZIkyZIk9SskSJIkyZIkSZIkSZIkSVL/RIIk
SZIkyZIkSZIkSZIkSfpFJEiSJEkyJEiCJEkyjAQSJEkyxJAgSZIkyZIk9SskSJIkyZIkSZIk
SZIkSVL/RIIkSZIkyZIkSZIkSZIkSfpFJEiSJEkyJEiCJEkyjAQSJEkyxJAgSZIkyZIk9Ssk
SJIkyZIkSZIkSZIkSVL/RIIkSZIkyZIkSZIkSZIkSfpFJEiSJEkyJEiCJEkyjAQSJEkyxJAg
SZIkyZIk9SskSJIkyZIkSZIkSZIkSVL/RIIkSZIkyZIkSZIkSZIkSfpFJEiSJEkyJEiCJEky
jAQSJEkyxJAgSZIkyZIk9SskSJIkyZIkSZIkSZIkSVL/RIIkSZIkyZIkSZIkSZIkSfpFJEiS
JEkyJEiCJEkyjAQSJEkyxJAgSZIkyZIk9SskSJIkyZIkSZIkSZIkSVL/RIIkSZIkyZIkSZIk
SZIkSfpFJEiSJEkyJEiCJEkyjAQSJEkyxJAgSZIkyZIk9SskSJIkyZIkSZIkSZIkSVL/RIIk
SZIkyZIkSZIkSZIkSfpFJEiSJEkyJEiCJEkyjAQSJEkyxJAgSZIkyZIk9SskSJIkyZIkSZIk
SZIkSVL/RIIkSZIkyZIkSZIkSZIkSfpFJEiSJEkyJEiCJEkyjAQSJEkyxJAgSZIkyZIk9Ssk
SJIkyZIkSZIkSZIkSVL/RIIkSZIkyZIkSZIkSZIkSfpFJEiSJEkyJEiCJEkyjAQSJEkyxJAg
SZIkyZIk9SskSJIkyZIkSZIkSZIkSVL/RIIkSZIkyZIkSZIkSZIkSfpFJEiSJEkyJEiCJEky
jAQSJEkyxJAgSZIkyZIk9SskSJIkyZIkSZIkSZIkSVL/RIIkSZIkyZIkSZIkSZIkSfpFJEiS
JEkyJEiCJEkyjAQSJEkyxJAgSZIkyZIk9SskSJIkyZIkSZIkSZIkSVL/RIIkSZIkyZIkSZIk
SZIkSfpFJEiSJEkyJEiCJEkyjAQSJEkyxJAgSZIkyZIk9SskSJIkyZIkSZIkSZIkSVL/RIIk
SZIkyZIkSZIkSZIkSfpFJEiSJEkyJEiCJEkyjAQSJEkyxJAgSZIkyZIk9SskSJIkyZIkSZIk
SZIkSVL/RIIkSZIkyZIkSZIkSZIkSfpFJEiSJEkyJEiCJEkyjAQSJEkyxJAgSZIkyZIk9Ssk
SJIkyZIkSZIkSZIkSVL/RIIkSZIkyZIkSZIkSZIkSfpFJEiSJEkyJEiCJEkyjAQSJEkyxJAg
SZIkyZIk9SskSJIkyZIkSZIkSZIkSVL/RIIkSZIkyZIkSZIkSZIkSfpFJEiSJEkyJEiCJEky
jAQSJEkyxJAgSZIkyZIk9SskSJIkyZIkSZIkSZIkSVL/RIIkSZIkyZIkSZIkSZIkSfpFJEiS
JEkyJEiCJEkyjAQSJEkyxJAgSZIkyZIk9SskSJIkyZIkSZIkSZIkSVL/RIIkSZIkyZIkSZIk
SZIkSfpFJEiSJEkyJEiCJEkyjAQSJEkyxJAgSZIkyZIk9SskSJIkyZIkSZIkSZIkSVL/RIIk
SZIkyZIkSZIkSZIkSfpFJEiSJEkyJEiCJEkyjAQSJEkyxJAgSZIkyZIk9SskSJIkyZIkSZIk
SZIkSVL/RIIkSZIkyZIkSZIkSZIkSfpFJEiSJEkyJEiCJEkyjAQSJEkyxJAgSZIkyZIk9Ssk
SJIkyZIkSZIkSZIkSVL/RIIkSZIkyZIkSZIkSZIkSfpFJEiSJEkyJEiCJEkyjAQSJEkyxJAg
SZIkyZIk9SskSJIkyZIkSZIkSZIkSVL/RIIk
"""

# --- KELAS UTAMA APLIKASI GUI ---
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Update Massal Marketplace")
        self.state('zoomed')

        # --- [SOLUSI IKON FINAL] ---
        # 1. Atur App ID unik agar Windows mengenali aplikasi ini
        myappid = 'fahmi.updatemassal.gui.2.6'
        if sys.platform == 'win32':
            try:
                ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)
            except Exception as e:
                print(f"Peringatan: Gagal mengatur App ID: {e}")

        # 2. Atur ikon jendela menggunakan iconbitmap() yang lebih andal
        try:
            # Gunakan fungsi resource_path untuk menemukan ikon
            icon_path = resource_path("logo.ico")
            self.iconbitmap(default=icon_path)
        except Exception as e:
            print(f"Error fatal: Tidak dapat memuat file ikon 'logo.ico'. Error: {e}")
        
        # 3. Muat logo untuk header (bisa dari Base64)
        try:
            self.logo_photo = tk.PhotoImage(data=LOGO_BASE64)
        except Exception as e:
            print(f"Error memuat logo header: {e}")
            self.logo_photo = None
        # --------------------------------

        # Konfigurasi Tema
        self.BG_COLOR = "#2E2E2E"
        self.FG_COLOR = "#FFFFFF"
        self.FRAME_COLOR = "#3C3C3C"
        self.BUTTON_COLOR = "#0078D7"
        self.BUTTON_ACTIVE_COLOR = "#005A9E"
        self.TEXT_COLOR = "#CCCCCC"
        self.LOG_BG_COLOR = "#1E1E1E"
        
        self.style = ttk.Style(self)
        self.style.theme_use('clam')
        self.configure(background=self.BG_COLOR)
        self._configure_styles()

        # Variabel aplikasi
        self.master_file = tk.StringVar()
        self.tiktok_files = []
        self.shopee_files = []
        self.log_queue = Queue()

        self._create_widgets()
        self.after(100, self._process_log_queue)

    def _configure_styles(self):
        self.style.configure('.', background=self.BG_COLOR, foreground=self.FG_COLOR, font=('Segoe UI', 10))
        self.style.configure('TFrame', background=self.BG_COLOR)
        self.style.configure('TLabel', background=self.BG_COLOR, foreground=self.FG_COLOR)
        self.style.configure('Header.TLabel', font=('Segoe UI', 14, 'bold'))
        self.style.configure('Footer.TLabel', font=('Segoe UI', 9))
        self.style.configure('TButton', padding=8, relief="flat", background=self.BUTTON_COLOR, foreground="white", font=('Segoe UI', 10, 'bold'))
        self.style.map('TButton', background=[('active', self.BUTTON_ACTIVE_COLOR)])
        self.style.configure('TLabelframe', background=self.BG_COLOR, bordercolor=self.FG_COLOR)
        self.style.configure('TLabelframe.Label', foreground=self.FG_COLOR, background=self.BG_COLOR, font=('Segoe UI', 11, 'bold'))

    def _create_widgets(self):
        main_frame = ttk.Frame(self)
        main_frame.pack(fill=tk.BOTH, expand=True)
        header_frame = ttk.Frame(main_frame, padding=(20, 10))
        header_frame.pack(side=tk.TOP, fill=tk.X)
        if self.logo_photo:
            small_logo = self.logo_photo.subsample(4, 4)
            logo_label = ttk.Label(header_frame, image=small_logo)
            logo_label.image = small_logo
            logo_label.pack(side=tk.LEFT, padx=(0, 15), anchor='w')
        ttk.Label(header_frame, text="Update Massal (Stock & Harga)", style='Header.TLabel', font=('Segoe UI', 20, 'bold')).pack(side=tk.LEFT, anchor='w')
        footer_frame = ttk.Frame(main_frame, padding=(10, 5))
        footer_frame.pack(side=tk.BOTTOM, fill=tk.X)
        ttk.Separator(footer_frame, orient='horizontal').pack(fill='x', pady=5)
        credit_label = ttk.Label(footer_frame, text="@khairudinfahmi", foreground="#4a90e2", cursor="hand2", style="Footer.TLabel")
        credit_label.pack()
        credit_label.bind("<Button-1>", lambda e: webbrowser.open_new("https://www.instagram.com/khairudinfahmi/"))
        content_frame = ttk.Frame(main_frame, padding=(20, 10))
        content_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        content_frame.columnconfigure(0, weight=1)
        content_frame.rowconfigure(1, weight=1)
        selection_frame = ttk.Frame(content_frame)
        selection_frame.grid(row=0, column=0, sticky="ew", pady=(0, 20))
        selection_frame.columnconfigure(1, weight=1)
        master_lf = ttk.LabelFrame(selection_frame, text="1. Pilih File Master Data")
        master_lf.grid(row=0, column=0, columnspan=2, sticky="ew", pady=5)
        master_lf.columnconfigure(1, weight=1)
        ttk.Button(master_lf, text="Pilih File...", command=self._select_master_file).grid(row=0, column=0, sticky='w', padx=10, pady=10)
        self.master_label = ttk.Label(master_lf, text="Belum ada file dipilih", wraplength=500, foreground=self.TEXT_COLOR)
        self.master_label.grid(row=0, column=1, sticky='w', padx=10)
        template_lf = ttk.LabelFrame(selection_frame, text="2. Pilih File Template")
        template_lf.grid(row=1, column=0, columnspan=2, sticky="ew", pady=5)
        template_lf.columnconfigure(1, weight=1)
        template_lf.columnconfigure(3, weight=1)
        ttk.Button(template_lf, text="Pilih File TikTok...", command=self._select_tiktok_files).grid(row=0, column=0, sticky='w', padx=10, pady=10)
        self.tiktok_list = tk.Listbox(template_lf, height=5, width=45, bg=self.FRAME_COLOR, fg=self.FG_COLOR, selectbackground=self.BUTTON_COLOR, borderwidth=0)
        self.tiktok_list.grid(row=0, column=1, sticky='ew', padx=10)
        ttk.Button(template_lf, text="Pilih File Shopee...", command=self._select_shopee_files).grid(row=0, column=2, sticky='w', padx=(20, 10), pady=10)
        self.shopee_list = tk.Listbox(template_lf, height=5, width=45, bg=self.FRAME_COLOR, fg=self.FG_COLOR, selectbackground=self.BUTTON_COLOR, borderwidth=0)
        self.shopee_list.grid(row=0, column=3, sticky='ew', padx=10)
        log_lf = ttk.LabelFrame(content_frame, text="Laporan Proses")
        log_lf.grid(row=1, column=0, sticky="nsew")
        log_lf.rowconfigure(0, weight=1)
        log_lf.columnconfigure(0, weight=1)
        self.log_area = scrolledtext.ScrolledText(log_lf, wrap=tk.WORD, font=('Consolas', 10), bg=self.LOG_BG_COLOR, fg=self.TEXT_COLOR, relief="flat")
        self.log_area.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)
        self.log_area.configure(state='disabled')
        self.run_button = ttk.Button(content_frame, text="🚀 Mulai Proses Update", command=self._start_processing_thread, style='TButton', padding=12)
        self.run_button.grid(row=2, column=0, pady=(20, 10))

    def _select_master_file(self):
        path = filedialog.askopenfilename(title="Pilih File Data Terkini", filetypes=[("File Excel", "*.xlsx"), ("Semua File", "*.*")])
        if path:
            self.master_file.set(path)
            self.master_label.config(text=os.path.basename(path))

    def _select_tiktok_files(self):
        paths = filedialog.askopenfilenames(title="Pilih File Template TikTok", filetypes=[("File Excel", "*.xlsx"), ("Semua File", "*.*")])
        if paths:
            self.tiktok_files = list(paths)
            self.tiktok_list.delete(0, tk.END)
            for path in self.tiktok_files:
                self.tiktok_list.insert(tk.END, os.path.basename(path))

    def _select_shopee_files(self):
        paths = filedialog.askopenfilenames(title="Pilih File Template Shopee", filetypes=[("File Excel", "*.xlsx"), ("Semua File", "*.*")])
        if paths:
            self.shopee_files = list(paths)
            self.shopee_list.delete(0, tk.END)
            for path in self.shopee_files:
                self.shopee_list.insert(tk.END, os.path.basename(path))

    def _log(self, message):
        self.log_queue.put(message)

    def _process_log_queue(self):
        while not self.log_queue.empty():
            message = self.log_queue.get_nowait()
            self.log_area.configure(state='normal')
            self.log_area.insert(tk.END, message + '\n')
            self.log_area.see(tk.END)
            self.log_area.configure(state='disabled')
        self.after(100, self._process_log_queue)

    def _start_processing_thread(self):
        if not self.master_file.get():
            messagebox.showerror("Error", "File Master belum dipilih!")
            return
        template_files = self.tiktok_files + self.shopee_files
        if not template_files:
            messagebox.showerror("Error", "Belum ada file template (TikTok/Shopee) yang dipilih!")
            return
        self.run_button.config(state='disabled', text="Sedang Memproses...")
        self.log_area.configure(state='normal')
        self.log_area.delete('1.0', tk.END)
        self.log_area.configure(state='disabled')
        processing_thread = threading.Thread(
            target=run_backend_process, 
            args=(self.master_file.get(), template_files, self._log, self._on_processing_complete)
        )
        processing_thread.daemon = True
        processing_thread.start()

    def _on_processing_complete(self):
        self.run_button.config(state='normal', text="🚀 Mulai Proses Update")
        messagebox.showinfo("Selesai", "Semua proses telah selesai!")

# --- FUNGSI-FUNGSI BACKEND ---
def clean_value(value):
    if pd.isna(value) or value == '': return ""
    cleaned_str = str(value).strip().upper().replace("O", "0")
    if cleaned_str.endswith('.0'): cleaned_str = cleaned_str[:-2]
    return cleaned_str
def save_dataframe_to_excel(df, original_filename, prefix, sheet_name_override, log_func):
    if df.empty: return True
    try:
        dir_name = os.path.dirname(original_filename) if os.path.dirname(original_filename) else '.'
        base_name = os.path.basename(original_filename)
        new_filename = os.path.join(dir_name, f"{prefix}_{base_name}")
        sheet_name = sheet_name_override or ('Sheet1' if 'sc' in base_name.lower() else 'Template')
        df.to_excel(new_filename, index=False, sheet_name=sheet_name)
        log_func(f"     💾 File '{prefix}' disimpan: '{os.path.basename(new_filename)}'")
        return True
    except Exception as e:
        log_func(f"     ❌ GAGAL menyimpan file '{os.path.basename(new_filename)}'. Error: {e}")
        return False
def find_header_row(file_path, sheet_name, keyword):
    try:
        df_peek = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=20)
        for i, row in df_peek.iterrows():
            if any(keyword.lower() in str(cell).lower() for cell in row.values):
                return i
    except Exception: pass
    return 0
def get_platform_config(df):
    cols_lower = {str(col).lower().strip(): str(col) for col in df.columns}
    if 'product_name' in cols_lower and 'seller_sku' in cols_lower:
        stock_col = next((col for col in df.columns if str(col).lower().startswith('warehouse_quantity')), None)
        if stock_col and 'price' in cols_lower:
            return 'tiktok', {'nama': cols_lower['product_name'], 'sku': cols_lower['seller_sku'], 'harga': cols_lower['price'], 'stok': stock_col}
    if 'nama produk' in cols_lower and 'sku induk' in cols_lower and 'sku' in cols_lower:
        if 'harga' in cols_lower and 'stok' in cols_lower:
            return 'shopee', {'nama': cols_lower['nama produk'], 'sku_induk': cols_lower['sku induk'], 'sku_utama': cols_lower['sku'], 'harga': cols_lower['harga'], 'stok': cols_lower['stok']}
    return 'unknown', None
def run_backend_process(master_file, template_files, log_func, on_complete_callback):
    try:
        log_func("--- MEMPROSES DATA MASTER ---")
        master_df = pd.read_excel(master_file, engine='openpyxl')
        master_search_map = {'kode': ['kodebarang'], 'barcode': ['barcode'], 'harga': ['hargajual'], 'stok': ['stok', 'stock']}
        master_cols = {k: next((c for t in v for c in master_df.columns if t in c.lower()), None) for k, v in master_search_map.items()}
        if not all(master_cols.values()):
            log_func("❌ GAGAL: Kolom penting di file master tidak ditemukan."); on_complete_callback(); return
        log_func(f"   - Kolom master terdeteksi: {master_cols}")
        master_df['kode_cleaned'] = master_df[master_cols['kode']].apply(clean_value)
        master_df['barcode_cleaned'] = master_df[master_cols['barcode']].apply(clean_value)
        lookup_map = {**pd.Series(list(zip(master_df[master_cols['harga']], master_df[master_cols['stok']])), index=master_df['kode_cleaned']).to_dict(), **pd.Series(list(zip(master_df[master_cols['harga']], master_df[master_cols['stok']])), index=master_df['barcode_cleaned']).to_dict()}
        lookup_map.pop("", None)
        log_func("   - Peta data master berhasil dibuat.")
        log_func("\n--- MEMPERBARUI SEMUA FILE TEMPLATE YANG DIPILIH ---")
        overall_summary = []
        for file in template_files:
            log_func(f"\n-------------------------------------------------------------")
            log_func(f"⚙️   Memproses file: '{os.path.basename(file)}'")
            try:
                xls = pd.ExcelFile(file)
                sheet_name = 'Template' if 'Template' in xls.sheet_names else xls.sheet_names[0]
                header_keyword = 'product_name' if 'tik' in os.path.basename(file).lower() else 'nama produk'
                header_row_index = find_header_row(file, sheet_name, header_keyword)
                template_df_raw = pd.read_excel(file, sheet_name=sheet_name, header=header_row_index)
                platform, template_cols = get_platform_config(template_df_raw)
                if platform == 'unknown':
                    overall_summary.append(f"'{os.path.basename(file)}': GAGAL (Platform/kolom tidak dikenali)"); continue
                template_df_cleaned = template_df_raw.dropna(subset=[template_cols['nama']]).copy()
                if platform == 'shopee':
                    sku_cols = [template_cols['sku_utama'], template_cols['sku_induk']]
                else: 
                    sku_cols = [template_cols['sku']]
                has_sku_mask = template_df_cleaned[sku_cols].notna().any(axis=1)
                template_df_to_process, empty_sku_rows = template_df_cleaned[has_sku_mask].copy(), template_df_cleaned[~has_sku_mask].copy()
                failed_rows_data, updated_rows_info = [], []
                for index, row in template_df_to_process.iterrows():
                    found_match, sku_to_find = False, None
                    if platform == 'tiktok':
                        sku_to_find = clean_value(row[template_cols['sku']])
                        if sku_to_find and sku_to_find in lookup_map: found_match = True
                    elif platform == 'shopee':
                        sku_utama, sku_induk = clean_value(row[template_cols['sku_utama']]), clean_value(row[template_cols['sku_induk']])
                        if sku_utama and sku_utama in lookup_map:
                            sku_to_find, found_match = sku_utama, True
                        elif sku_induk and sku_induk in lookup_map:
                            sku_to_find, found_match = sku_induk, True
                    if found_match:
                        new_price, new_stock = lookup_map[sku_to_find]
                        updated_rows_info.append({'original_index': index, 'harga': new_price, 'stok': max(0, int(float(new_stock)))})
                    else:
                        failed_rows_data.append(row.to_dict())
                dir_name, base_name = os.path.dirname(file) or '.', os.path.basename(file)
                save_dataframe_to_excel(pd.DataFrame(failed_rows_data), file, "GAGAL", "Produk Gagal Ditemukan", log_func)
                save_dataframe_to_excel(empty_sku_rows, file, "KOSONG", "Produk SKU Kosong", log_func)
                if updated_rows_info:
                    workbook = openpyxl.load_workbook(file)
                    worksheet = workbook[sheet_name]
                    col_map = {str(cell.value): cell.column for cell in worksheet[header_row_index + 1]}
                    harga_col_idx = col_map.get(template_cols['harga'])
                    stok_col_idx = col_map.get(template_cols['stok'])
                    if harga_col_idx and stok_col_idx:
                        for update_info in updated_rows_info:
                            excel_row = update_info['original_index'] + header_row_index + 2
                            worksheet.cell(row=excel_row, column=harga_col_idx).value = update_info['harga']
                            worksheet.cell(row=excel_row, column=stok_col_idx).value = update_info['stok']
                    update_filename = os.path.join(dir_name, f"UPDATE_{base_name}")
                    workbook.save(update_filename)
                    log_func(f"     💾 File 'UPDATE' disimpan: '{os.path.basename(update_filename)}'")
                else:
                    log_func("     ℹ️ Tidak ada data yang cocok untuk diperbarui di file ini.")
                summary_text = (f"'{base_name}' ({platform.title()}):\n"
                                f"         - Total Produk di File : {len(template_df_cleaned)}\n"
                                f"         - Produk SKU Kosong    : {len(empty_sku_rows)}\n"
                                f"         - Berhasil Update      : {len(updated_rows_info)}\n"
                                f"         - Gagal Ditemukan      : {len(failed_rows_data)}")
                overall_summary.append(summary_text)
            except Exception as e:
                error_msg = f"'{os.path.basename(file)}': PROSES ERROR ({e})"
                log_func(f"     ❌ {error_msg}")
                overall_summary.append(error_msg)
        log_func("\n" + "="*70)
        log_func("                       LAPORAN HASIL PEMBARUAN")
        log_func("="*70)
        for summary in overall_summary: log_func(summary)
    except Exception as e:
        log_func(f"\n\n❌❌❌ TERJADI ERROR FATAL PADA PROGRAM ❌❌❌\nError: {e}")
    finally:
        on_complete_callback()

# --- Titik Masuk Program ---
if __name__ == "__main__":
    app = App()
    app.mainloop()
